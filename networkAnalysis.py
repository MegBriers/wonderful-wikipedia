# -*- coding: utf-8 -*-
"""
Created on Wed Feb 23 09:09:30 2022

CODE TO PERFORM ANALYSIS OF POPULAR FIGURES ACROSS
WIKIPEDIA ARTICLES AND THE EPSILON DATA WORK

@author: Meg
"""

import helper
import os
import Levenshtein
from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import csv
import sys

# can be uncommented as required (changes properties of plt frame)
# plt.rc('figure', figsize=(25, 20))
# plt.rcParams.update({'font.size': 29})
my_colors = ['#79CDCD', '#FF7F00']
folders = ['maths', 'philosophy']


def update_counts(figures, people, cur_person):
    """

    A method to update the record of people who have been mentioned
    across multiple Wikipedia articles
    each line of the figures dictionary will store a mentioned figure, as well
    as the number of times they have been mentioned, and who they have been
    mentioned by

    Parameters
    ----------
    figures : dictionary
        stores the people mentioned in articles, how many times they have been mentioned and who they have been mentioned by
        STRUCTURE -> figure : [no_of_mentions, [people mentioned]]

    people : array of strings
        all the unique people mentioned in the article of cur_person

    cur_person : string
        the person whose article mentions we are working with currently

    Returns
    -------
    figures : dictionary
        as above, just updated for new counts


    """
    for peep in people:
        if peep in figures.keys():
            cur_values = figures.get(peep, 0)
            figures[peep] = [cur_values[0] + 1, cur_values[1] + [cur_person]]
        else:
            figures[peep] = [1, [cur_person]]
    return figures


def setup():
    """

    A method that creates files to hold the data generated from the run of the code on the
    whole network of philosophers and mathematicians

    The files store information such as the category that the subject of the
    article falls under, their gender, and how many mentions were picked up using
    a specified method

    Each person will be in the mentions file twice (one for the record with spacy extraction,
    and one with the wikidata extraction)

    Also creates a file to store how often each person found in these articles has been
    mentioned/linked across the network

    Parameters
    ----------
    None.

    Returns
    -------
    None.


    """
    subfolder = {"https://en.wikipedia.org/wiki/Category:19th-century_British_philosophers": "philosophy",
                 "https://en.wikipedia.org/wiki/Category:19th-century_British_mathematicians": "maths"}

    # these will be used to store the number of times people mentioned in a page are mentioned across the whole network
    pop = {'spacy': [], 'wikidata': []}
    maths_pop_s = {}
    phil_pop_s = {}
    maths_pop_w = {}
    phil_pop_w = {}

    with open('./analysis/mentions.csv', 'w') as f:
        f.write('name,method,category,mentions,length,gender')
        f.write('\n')
        for category in subfolder.keys():
            list_of_people = helper.get_list(category)
            for person in list_of_people:
                person_format = person.replace(" ", "_")

                if "," in person:
                    split = person.split(",", 1)
                    person = split[0]

                # finding the relevant path files for the person's mentions
                file_path_linked = helper.get_file_path(person_format + "_Linked.txt", "\output\wikidata\\network")
                file_path_unlinked = helper.get_file_path(person_format + "_Unlinked.txt", "\output\spacy")

                # if something has gone wrong during name extraction
                if file_path_linked is None or file_path_unlinked is None:
                    continue

                file_linked = open(file_path_linked)
                file_unlinked = open(file_path_unlinked)

                # get rid of the first line as just stores length of file
                length_of_file = file_unlinked.readline().rstrip()
                file_linked.readline().rstrip()

                mentioned = list(set(line.rstrip("\n") for count, line in enumerate(file_unlinked)))
                linked = list(set(line.rstrip("\n") for count, line in enumerate(file_linked)))

                # dealing with requests that went wrong (not counted towards analysis because it will throw off figures)
                if len(mentioned) == 0 or len(linked) == 0:
                    continue

                # updating counts for mentions of each person in each category
                if subfolder[category] == "maths":
                    maths_pop_s = update_counts(maths_pop_s, mentioned, person)
                    maths_pop_w = update_counts(maths_pop_w, linked, person)
                else:
                    phil_pop_s = update_counts(phil_pop_s, mentioned, person)
                    phil_pop_w = update_counts(phil_pop_w, linked, person)

                # longest identifiable substring to decide on gender
                if 'female' in file_path_linked:
                    gender = 'female'
                else:
                    gender = 'male'

                f.write(person + ",spacy" + "," + subfolder[category] + "," + str(len(mentioned)) + "," + str(
                    length_of_file) + "," + gender)
                f.write('\n')
                f.write(person + ",wikidata" + "," + subfolder[category] + "," + str(len(linked)) + "," + str(
                    length_of_file) + "," + gender)
                f.write('\n')
                file_linked.close()
                file_unlinked.close()
    f.close()

    pop['spacy'] = {'maths': maths_pop_s, 'phil': phil_pop_s}
    pop['wikidata'] = {'maths': maths_pop_w, 'phil': phil_pop_w}

    # writing the popular figures to a file
    with open('./analysis/popular_figures.txt', 'w') as f:
        for method in pop.keys():
            for group in pop[method].keys():
                for person in pop[method][group].keys():
                    first_name = person
                    if "," in person:
                        split = person.split(",", 1)
                        first_name = split[0]
                    f.write(
                        method + ',' + group + ',' + first_name + "," + str(pop[method][group][person][0]) + "," + str(
                            pop[method][group][person][1]))
                    f.write('\n')
    f.close()


def popular_figs():
    """

     A method that identifies the most commonly mentioned/linked figures
     across the networks of mathematicians and philosophers and produces
     graphs illustrating the top 10 most linked to figures

     Parameters
     ----------
     None.

     Returns
     -------
     None.


     """

    pop_figs = [{}, {}, {}, {}]

    # each row stores the following information in the indices
    # 0 - method used to extract name
    # 1 - category that the mentions come from
    # 2 - name being mentioned
    # 3 - how many wikipedia articles mentioned them
    # 4 - which wikipedia articles mentioned them

    # ideally needs to be shifted to have header values so it can be accessed by column title
    # which would allow more flexibility with order of file
    with open('./analysis/popular_figures.txt') as csvfile:
        current_reader = csv.reader(csvfile)
        for row in current_reader:
            name = row[2]
            # not required for the analysis as identified error in wikidata extraction
            if "content" in row[2]:
                continue

            if row[0] == "spacy":
                if row[1] == "maths":
                    pop_figs[0][name] = int(row[3])
                else:
                    pop_figs[1][name] = int(row[3])
            else:
                if row[1] == "maths":
                    pop_figs[2][name] = int(row[3])
                else:
                    pop_figs[3][name] = int(row[3])

    csvfile.close()

    # generates four graphs in this order; maths spacy, maths wikidata, philosophy spacy, philosophy wikidata
    for graph in pop_figs:
        sorted_list = dict(sorted(graph.items(), key=lambda item: item[1], reverse=True))

        # only want top 10 values for the graph
        top = {key: sorted_list[key] for count, key in enumerate(sorted_list.keys()) if count < 10}

        # shortening the name that will be displayed on the graph if needed
        shortened_top = {}
        for person in top.keys():
            # so the full names can be identifiable
            print(person, helper.initials(person,14))
            shortened_top[helper.initials(person,14)] = top[person]

        fig, ax = plt.subplots()

        ax.barh(list(shortened_top.keys()), list(top.values()), color=my_colors[0])
        ax.invert_yaxis()
        ax.set_xlabel('Number of mentions')
        ax.set_xticks(np.arange(0, max(list(top.values())) + 1, 1))

        plt.show()


def epsilon_analysis():
    """

    Code that compares the unique people known to be in correspondence with
    Mary Somerville and those mentioned on her Wikipedia page


    Parameters
    ----------
    None.

    Returns
    -------
    None.

    """

    # needs refactoring

    stdoutOrigin = sys.stdout
    sys.stdout = open("analysis/epsilon.txt", "w", encoding="utf-8")

    wb = load_workbook('data\somerville_letters.xlsx')
    sheet_ranges = wb['Sheet1']

    row_count = sheet_ranges.max_row

    correspondences = {}

    for i in range(2, row_count):
        # the relevant places in the excel file
        person_from = 'D' + str(i)
        person_to = 'E' + str(i)

        person_from = sheet_ranges[person_from].value
        person_to = sheet_ranges[person_to].value

        # either writing to or from her or her husband
        # getting rid of family members in the analysis
        if 'Somerville' in str(person_from):
            if not (person_to in correspondences.keys()):
                correspondences[person_to] = 1
            else:
                correspondences[person_to] += 1
        else:
            if 'Somerville' in str(person_to):
                if not (person_from in correspondences):
                    correspondences[person_from] = 1
                else:
                    correspondences[person_from] += 1

    print("The name of people in correspondence with Somerville:")
    print(correspondences)
    print(".・。.・゜✭・.・✫・゜・。.")
    print("Number of unique correspondences:")
    print(len(correspondences))
    print(".・。.・゜✭・.・✫・゜・。.")

    complete = []
    # the list of people manually identified from the somerville page
    manual = pd.read_csv("./people/Mary_Somerville.txt")

    # dealing with the way that file is set up
    for index, row in manual.iterrows():
        complete.append(row["Target"])

    print("Names manually identified from Somerville page:")
    print(complete)
    print(".・。.・゜✭・.・✫・゜・。.")
    print(len(complete))

    common_names = []

    for name in correspondences.keys():
        found = False
        for line in complete:
            if not found:
                # experimentation with 0.7 meant too many false positives were getting through. could also add in a
                # check for Ada Byron in Ada Byron (King) but then run the risk of combining fathers and sons if they
                # are just marked different by Jnr. at > 0.8 we get rid of a lot of misidentified williams but we also
                # lose the Humboldt being identified by his proper title
                if Levenshtein.ratio(name, line) > .8 and not (line in common_names):
                    print("")
                    print("Identified correspondence:")
                    print(name)
                    print("Identified in article:")
                    print(line)
                    print("Levenshtein ratio:")
                    print(Levenshtein.ratio(name, line))
                    common_names.append(line)
                    found = True

    print(".・。.・゜✭・.・✫・゜・。.")
    print("Number of people found in correspondence and Wikipedia")
    print(len(common_names))
    print("People identified in correspondence and Wikipedia")
    print(common_names)
    print(".・。.・゜✭・.・✫・゜・。.")
    print("People written to in order of decreasing number of correspondences")
    new = {k: v for k, v in sorted(correspondences.items(), key=lambda item: item[1], reverse=True)}
    print(new)

    sys.stdout.close()
    sys.stdout = stdoutOrigin


def start():
    """
    driver method that ensures the files are set up before continuing with analysis
    """
    if not (os.path.isfile('./analysis/popular_figures.txt')) or not (os.path.isfile('./analysis/mentions.csv')):
        setup()

    popular_figs()
    epsilon_analysis()